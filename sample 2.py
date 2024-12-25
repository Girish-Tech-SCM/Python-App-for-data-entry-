import tkinter as tk
from tkinter import ttk, messagebox, filedialog 
import openpyxl

def save_data():
    """
    Handles the button click event:
    - Gets user input from the GUI.
    - Updates the Excel file.
    - Displays a success message.
    """
    try:
        name = name_entry.get()
        age = age_entry.get()
        gender = gender_combobox.get()
        education = education_entry.get()

        if not name or not age or not gender or not education:
            messagebox.showwarning("Warning", "Please fill all fields.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )

        if not file_path:
            return  # User canceled the file selection

        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # Add headers if not present. Also this is the line that gave the tuple error message
        if not any(cell.value for row in sheet['A1:D1'] for cell in row):
            sheet['A1'] = 'Name'
            sheet['B1'] = 'Age'
            sheet['C1'] = 'Gender'
            sheet['D1'] = 'Education'

        # Find the next available row
        next_row = sheet.max_row + 1

        # Write data to the sheet
        sheet.cell(row=next_row, column=1, value=name)
        sheet.cell(row=next_row, column=2, value=age)
        sheet.cell(row=next_row, column=3, value=gender)
        sheet.cell(row=next_row, column=4, value=education)

        # Save the workbook
        workbook.save(file_path)

        messagebox.showinfo("Success", "Data saved successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Create the main window
root = tk.Tk()
root.title("Data Entry Application")

# Create GUI elements
name_label = ttk.Label(root, text="Name:")
name_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
name_entry = ttk.Entry(root, width=30)
name_entry.grid(row=0, column=1, padx=10, pady=5)

age_label = ttk.Label(root, text="Age:")
age_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
age_entry = ttk.Entry(root, width=10)
age_entry.grid(row=1, column=1, padx=10, pady=5)

gender_label = ttk.Label(root, text="Gender:")
gender_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
gender_combobox = ttk.Combobox(root, values=["Male", "Female", "Other"], state="readonly")
gender_combobox.grid(row=2, column=1, padx=10, pady=5)
gender_combobox.current(0)  # Set default value

education_label = ttk.Label(root, text="Education:")
education_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")
education_entry = ttk.Entry(root, width=30)
education_entry.grid(row=3, column=1, padx=10, pady=5)

save_button = ttk.Button(root, text="Save", command=save_data)
save_button.grid(row=4, column=1, padx=10, pady=10)

root.mainloop()